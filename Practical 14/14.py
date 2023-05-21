import xml.etree.ElementTree as ET
import openpyxl

# Load the XML file
tree = ET.parse('go_obo.xml')
root = tree.getroot()

# Create a new Excel workbook and worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Write the headers to the worksheet
worksheet['A1'] = 'GO id'
worksheet['B1'] = 'Term name'
worksheet['C1'] = 'Definition string'
worksheet['D1'] = 'Number of child nodes'

# Find all occurrences of "autophagosome" in the <defstr> element
autophagosome_terms = []
for term in root.findall('term'):
    defstr = term.find('def/defstr').text
    if 'autophagosome' in defstr:
        go_id = term.find('id').text
        term_name = term.find('name').text
        definition_string = defstr
        autophagosome_terms.append((go_id, term_name, definition_string))

# Find the number of child nodes for each autophagosome-related term
for i, term in enumerate(autophagosome_terms):
    go_id = term[0]
    num_child_nodes = 0
    for child in root.findall('term/[is_a="' + go_id + '"]'):
        num_child_nodes += 1
        queue = [child]
        while queue:
            node = queue.pop(0)
            num_child_nodes += len(node.findall('is_a'))
            queue.extend(root.findall('term/[is_a="' + node.find('id').text + '"]'))
    worksheet.cell(row=i+2, column=1, value=go_id)
    worksheet.cell(row=i+2, column=2, value=term[1])
    worksheet.cell(row=i+2, column=3, value=term[2])
    worksheet.cell(row=i+2, column=4, value=num_child_nodes)

# Save the workbook
workbook.save('autophagosome.xlsx')